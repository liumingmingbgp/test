import openpyxl
wb = openpyxl.Workbook()
ws = wb.active
print('ws title:', ws.title)
ws2 = wb.create_sheet("NewTitle", 1)
ws2.title = 'MySheet'
# ws2.cell(row=2, column=2).value = '二两'
# wb.save('test2.xlsx')

for row in ws2['A2':'C3']:
    for cell in row:
        cell.value='haha'

        wb.save('test2.xlsx')
print(ws2.cell(row=3, column=3).value)

ws3 = wb.active
rows = [
    ['ID','name','age'],
    [1,'张三','28'],
    [2,'李四','25'],
    [3,'王五','40'],
    [4,'赵六','23'],
]
for row in rows:
    ws.append(row)
    wb.save('test2.xlsx')