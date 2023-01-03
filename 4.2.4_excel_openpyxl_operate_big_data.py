from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.comments import Comment
from openpyxl.styles import Font

wb = Workbook(write_only=True)
ws = wb.create_sheet()
cell = WriteOnlyCell(ws, value='write_only状态写入的内容')
cell.font = Font(name='微软雅黑', size=36)
cell.comment = Comment(text='这是一个批注', author='二两')
ws.append([cell, 2.333, None])
wb.save('write_only.xlsx')