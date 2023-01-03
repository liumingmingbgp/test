import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font
import calendar
calendar.setfirstweekday(firstweekday=6)
year = 2020
wb = openpyxl.Workbook()
# print(calendar.monthcalendar(2023,1))
for i in range(1, 13):
    sheet = wb.create_sheet(index=0, title=str(i)+'月')
    for j in range(len(calendar.monthcalendar(year, i))):
        for k in range(len(calendar.monthcalendar(year, i)[j])):
            value = calendar.monthcalendar(year, i)[j][k]
            if value == 0:
                value = '' # 将0变为空值，没有日期的单元格填空值
                sheet.cell(row=j+9, column=k+1).value = value
            else:
                sheet.cell(row=j+9, column=k+1).value = value
                sheet.cell(row=j+9, column=k+1).font = Font(u'微软雅黑', size=11)
    wb.save('calendar.xlsx')