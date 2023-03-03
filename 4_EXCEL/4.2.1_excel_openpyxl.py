import openpyxl
wb = openpyxl.load_workbook('test.xlsx')
ws = wb.worksheets[0]
print(ws.cell(row=3, column=2).value)
print('最大行数:',ws.max_row)
print('最小行数:',ws.min_row)
print('最大列数:',ws.max_column)
print('最小列数:',ws.min_column)
for col in ws.iter_cols(min_col=1, max_col=2, max_row=2, values_only=True):
    print(col)

all_values = ws.values
print(type(all_values))
for i, value in enumerate(all_values):
    print(value)
    if i==2:
        break