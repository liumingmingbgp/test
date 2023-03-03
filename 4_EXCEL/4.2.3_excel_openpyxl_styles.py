import openpyxl
from openpyxl.styles import Font, colors, PatternFill, Border, Side, Alignment
wb = openpyxl.Workbook()
ws = wb.active
rows = [
    ['ID','name','age', 'Gender'],
    [1,'张三','28', 'man'],
    [2,'李四','25', 'woman'],
    [3,'王五','40', 'man'],
    [4,'赵六','23', 'woman'],
]
for row in rows:
    ws.append(row)
wb.save('test3.xlsx')

font = Font(name='微软雅黑', size=25, italic=True, color=colors.BLUE, bold=True)
ws['A1'].font = font
fill = PatternFill(fill_type='solid', start_color=colors.BLUE)
ws['B1'].fill = fill
border = Border(
    left=Side(border_style='double', color='FFBB00'),
    right=Side(border_style='double', color='FFBB00'),
    top=Side(border_style='double', color='FFBB00'),
    bottom=Side(border_style='double', color='FFBB00'),
)
ws['C1'].border = border
align = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws['D1'].alignment = align
ws.row_dimensions[3].height = 40
ws.column_dimensions['A'].width = 30
ws.merge_cells('A7:C7')
ws.merge_cells('A9:C13')
ws['A9'] = '合并单元格'
wb.save('test3.xlsx')