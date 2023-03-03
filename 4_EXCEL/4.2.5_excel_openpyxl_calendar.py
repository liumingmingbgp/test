import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import calendar
calendar.setfirstweekday(firstweekday=6)
year = 2020
wb = openpyxl.Workbook()
# print(calendar.monthcalendar(2023,1))
for i in range(1, 13):
    sheet = wb.create_sheet(index=0, title=str(i)+'月')
    for j in range(len(calendar.monthcalendar(year, i))):
        for k in range(len(calendar.monthcalendar(year, i)[j])):
            value = calendar.monthcalendar(year, i)[j][k]
            if value == 0:
                value = '' # 将0变为空值，没有日期的单元格填空值
                sheet.cell(row=j+9, column=k+1).value = value
            else:
                sheet.cell(row=j+9, column=k+1).value = value
                sheet.cell(row=j+9, column=k+1).font = Font(u'微软雅黑', size=11)
    align = Alignment(horizontal='right', vertical='center')
    fill = PatternFill('solid', fgColor="99cccc")
    for k1 in range(1, 50):
        for k2 in range(1, 50):
            sheet.cell(row=k1, column=k2).fill = fill
    days = ['星期日', '星期一', '星期二', '星期三', '星期四', '星期五', '星期六']
    num = 0
    for k3 in range(1, 8):
        sheet.cell(row=8, column=k3).value = days[num]
        sheet.cell(row=8, column=k3).alignment = align
        sheet.cell(row=8, column=k3).font = Font(u'微软雅黑', size=11)
        c_char = get_column_letter(k3)
        sheet.column_dimensions[c_char].width = 12
        num += 1
    for k4 in range(8, 14):
        sheet.row_dimensions[k4].height = 30
    sheet.merge_cells('I1:P20')
    img = Image('ICON.jpeg')
    newsize = (300, 200)
    img.width, img.height = newsize
    sheet.add_image(img, 'I2')
    sheet.cell(row=3, column=1).value = f'{year}年'
    sheet.cell(row=4, column=1).value = str(i)+'月'
    sheet.cell(row=3, column=1).font = Font(u'微软雅黑', size=16, bold=True, color='FF7887')
    sheet.cell(row=4, column=1).font = Font(u'微软雅黑', size=16, bold=True, color='FF7887')
    sheet.cell(row=3, column=1).alignment = align
    sheet.cell(row=4, column=1).alignment = align
wb.save('calendar.xlsx')