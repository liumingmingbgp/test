import openpyxl
wb = openpyxl.Workbook()
ws = wb.active
print('ws title:', ws.title)
ws2 = wb.create_sheet("NewTitle", 1)
ws2.title = 'MySheet'
ws2.cell(row=2, column=2).value = '二两'
wb.save('test2.xlsx')